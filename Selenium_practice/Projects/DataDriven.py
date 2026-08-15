from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

# Load excel Sheet
workbook = load_workbook('Projects/data0.xlsx')

# Selecting Active Sheet
sheet = workbook.active

# =========================
# Chrome Setup
# =========================

options = webdriver.ChromeOptions()

# Disable Chrome's "Change your password" breach warning
options.add_experimental_option(
    "prefs",
    {
        "profile.password_manager_leak_detection": False
    }
)

driver = webdriver.Chrome(options=options)
driver.maximize_window()


for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
    username = row[0]
    password = row[1]

    driver.get("https://www.saucedemo.com/")
    time.sleep(5)

    driver.find_element(By.ID,"user-name").send_keys(username)
    driver.find_element(By.ID,"password").send_keys(password)
    driver.find_element(By.ID,"login-button").click()
    time.sleep(5)


    driver.find_element(By.XPATH,"//button[@id='react-burger-menu-btn']").click()
    time.sleep(5)

    driver.find_element(By.XPATH,"//a[@id='logout_sidebar_link']").click()
    time.sleep(5)

driver.quit()


